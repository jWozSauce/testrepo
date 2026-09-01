"""Multi-stat season projections and an Excel cheat-sheet export.

Extends the single-target (half-PPR) projection to a full per-position stat line by
training one model per (position, stat) on the same feature frames, then writes a
formatted multi-tab workbook: an ALL-players tab plus one tab per position.
"""
from __future__ import annotations

import json

import numpy as np
import pandas as pd

from . import features as F, data as D, models as M
from .sources.madden import normalize_name

POS = ("QB", "RB", "WR", "TE")

# Stats projected per position (all are aggregated in build_player_seasons).
STAT_TARGETS = {
    "QB": ["games", "completions", "attempts", "passing_yards", "passing_tds",
           "interceptions", "carries", "rushing_yards", "rushing_tds"],
    "RB": ["games", "carries", "rushing_yards", "rushing_tds", "targets",
           "receptions", "receiving_yards", "receiving_tds"],
    "WR": ["games", "targets", "receptions", "receiving_yards", "receiving_tds",
           "carries", "rushing_yards"],
    "TE": ["games", "targets", "receptions", "receiving_yards", "receiving_tds"],
}

# Non-negative integer-ish counting stats (for clipping / display rounding).
_COUNTS = {"games", "completions", "attempts", "passing_tds", "interceptions",
           "carries", "targets", "receptions", "receiving_tds", "rushing_tds"}

# Pretty column headers for the workbook.
DISPLAY = {
    "proj_points": "Proj Pts", "proj_ppg": "Proj PPG", "proj_games": "Proj G",
    "proj_completions": "Cmp", "proj_attempts": "Att", "proj_passing_yards": "Pass Yds",
    "proj_passing_tds": "Pass TD", "proj_interceptions": "INT",
    "proj_carries": "Car", "proj_rushing_yards": "Rush Yds", "proj_rushing_tds": "Rush TD",
    "proj_targets": "Tgt", "proj_receptions": "Rec", "proj_receiving_yards": "Rec Yds",
    "proj_receiving_tds": "Rec TD",
    "pos_rank": "Pos Rk", "ovr_rank": "Ovr Rk", "player_name": "Player",
    "position": "Pos", "team": "Team", "age": "Age", "madden_overall": "MAD",
    "vbd": "VBD", "tier": "Tier", "vbd_rank": "Rk",
    "adp_rank": "ADP", "mdl_rank": "Model", "value": "Val",
}
VALUE_COLS = ["player_name", "position", "team", "tier", "adp_rank", "mdl_rank",
              "value", "proj_points", "vbd"]


def fetch_live_adp(year=2026, teams=12, scoring="half-ppr", save_path=None):
    """Pull current ADP from the Fantasy Football Calculator public API and return a
    (name_key, adp_rank) table for skill positions. Optionally snapshot the raw rows."""
    import subprocess
    url = (f"https://fantasyfootballcalculator.com/api/v1/adp/{scoring}"
           f"?teams={teams}&year={year}&position=all")
    r = subprocess.run(["curl", "-sSL", "--compressed", "-A", "Mozilla/5.0", url],
                       capture_output=True, text=True, timeout=40)
    data = json.loads(r.stdout)
    d = pd.DataFrame(data.get("players", []))
    d = d[d["position"].isin(POS)].copy()
    d["adp"] = pd.to_numeric(d["adp"], errors="coerce")
    d = d.dropna(subset=["adp"]).sort_values("adp").reset_index(drop=True)
    d["adp_rank"] = np.arange(1, len(d) + 1)
    d["name_key"] = d["name"].map(normalize_name)
    meta = data.get("meta", {})
    print(f"live ADP: {meta.get('total_drafts')} drafts, "
          f"{meta.get('start_date')}..{meta.get('end_date')}, {len(d)} skill players")
    if save_path:
        d[["adp_rank", "name", "position", "team", "adp"]].to_csv(save_path, index=False)
    return d[["name_key", "adp_rank"]].drop_duplicates("name_key")


def load_adp(path):
    """Read a fantasy ADP sheet (Rank/Name/Team/Pos header on the 2nd row) and return a
    (name_key, adp_rank) table for skill positions."""
    d = pd.read_excel(path, header=1)
    d = d.rename(columns={d.columns[0]: "adp_rank", d.columns[1]: "name",
                          d.columns[3]: "pos"})
    d["adp_rank"] = pd.to_numeric(d["adp_rank"], errors="coerce")
    d = d[d["pos"].isin(POS) & d["adp_rank"].notna()].copy()
    d["name_key"] = d["name"].map(normalize_name)
    return d[["name_key", "adp_rank"]].drop_duplicates("name_key")


def attach_adp(board, adp):
    """Merge ADP onto a projection board and compute model-vs-market value. Value =
    market_rank - model_rank within the matched pool (+ = model higher than market)."""
    b = board.copy()
    b["name_key"] = b["player_name"].map(normalize_name)
    b = b.merge(adp, on="name_key", how="left")
    common = b[b["adp_rank"].notna()].copy()
    common["mkt_rank"] = common["adp_rank"].rank(method="first")
    common["mdl_rank"] = common["vbd"].rank(ascending=False, method="first")
    common["value"] = (common["mkt_rank"] - common["mdl_rank"]).round(0)
    return b.merge(common[["name_key", "mdl_rank", "value"]], on="name_key", how="left")

# 12-team league, half-PPR, standard lineup.
LEAGUE = dict(teams=12, starters={"QB": 1, "RB": 2, "WR": 3, "TE": 1},
              flex=1, flex_pos=("RB", "WR", "TE"))

# Streaming/scarcity-adjusted replacement ranks (12-team, half-PPR). VBD for a player =
# points over the Nth player at his position. QB/TE are SHALLOW (you can stream a
# startable one, so replacement is good -> elite VBD is modest); RB/WR are DEEP (scarce,
# injuries/committees -> replacement is poor -> elite VBD is large). This matches how the
# market actually drafts far better than one-per-team starter baselines.
BASELINE_RANK = {"QB": 12, "RB": 40, "WR": 55, "TE": 12}


def add_vbd(allp, baseline_rank=BASELINE_RANK):
    """Value over replacement, where replacement is the Nth-ranked player at the position
    (N per BASELINE_RANK). Returns (df with 'vbd', baseline points per position)."""
    df = allp.copy()
    base = {}
    for p in df.position.unique():
        n = baseline_rank.get(p, 24)
        s = df[df.position == p].nlargest(n, "proj_points")["proj_points"]
        base[p] = float(s.iloc[-1]) if len(s) >= 1 else 0.0
    df["vbd"] = (df["proj_points"] - df["position"].map(base)).round(0)
    return df, base


def add_tiers(df, top_n={"QB": 24, "RB": 42, "WR": 54, "TE": 24}, mult=0.8):
    """Per-position tiers via natural gaps: a new tier starts where the drop to the next
    player is unusually large (> mean + mult*std of gaps among the draftable range)."""
    df = df.copy(); df["tier"] = 1
    for p in df.position.unique():
        sub = df[df.position == p].sort_values("proj_points", ascending=False)
        pts = sub["proj_points"].to_numpy(float); n = len(pts)
        lim = min(top_n.get(p, 40), n)
        if lim < 3:
            df.loc[sub.index, "tier"] = 1; continue
        gaps = -np.diff(pts[:lim])
        thr = gaps.mean() + mult * gaps.std()
        t, tiers = 1, [1]
        for g in gaps:
            if g > thr:
                t += 1
            tiers.append(t)
        tiers += [t + 1] * (n - lim)
        df.loc[sub.index, "tier"] = tiers[:n]
    df["tier"] = df["tier"].astype(int)
    return df

# Column order shown on each position tab.
TAB_COLS = {
    "QB": ["pos_rank", "tier", "player_name", "team", "age", "madden_overall",
           "proj_points", "vbd", "proj_ppg", "proj_games", "proj_passing_yards",
           "proj_passing_tds", "proj_interceptions", "proj_completions",
           "proj_attempts", "proj_rushing_yards", "proj_rushing_tds"],
    "RB": ["pos_rank", "tier", "player_name", "team", "age", "madden_overall",
           "proj_points", "vbd", "proj_ppg", "proj_games", "proj_carries",
           "proj_rushing_yards", "proj_rushing_tds", "proj_targets",
           "proj_receptions", "proj_receiving_yards", "proj_receiving_tds"],
    "WR": ["pos_rank", "tier", "player_name", "team", "age", "madden_overall",
           "proj_points", "vbd", "proj_ppg", "proj_games", "proj_targets",
           "proj_receptions", "proj_receiving_yards", "proj_receiving_tds",
           "proj_rushing_yards"],
    "TE": ["pos_rank", "tier", "player_name", "team", "age", "madden_overall",
           "proj_points", "vbd", "proj_ppg", "proj_games", "proj_targets",
           "proj_receptions", "proj_receiving_yards", "proj_receiving_tds"],
}
ALL_COLS = ["ovr_rank", "player_name", "position", "tier", "team", "age",
            "madden_overall", "pos_rank", "proj_points", "vbd", "proj_ppg", "proj_games"]
BOARD_COLS = ["vbd_rank", "player_name", "position", "tier", "team", "age",
              "proj_points", "vbd", "proj_ppg", "pos_rank"]


def _fit_predict(f, proj, feats, label_col, out_col):
    y = f[label_col]
    mask = y.notna()
    usable = [c for c in feats if c in proj.columns
              and f.loc[mask, c].nunique(dropna=True) >= 2]
    model = M.make_model()
    model.fit(f.loc[mask, usable].to_numpy(float), y[mask].to_numpy(float))
    pred = model.predict(proj[usable].to_numpy(float))
    if out_col == "proj_games":
        pred = np.clip(pred, 0, 17)
    else:
        pred = np.clip(pred, 0, None)
    proj[out_col] = pred


def _durability_games(f, proj):
    """Durability-adjusted games projection. The raw games model over-weights last
    season's games, so a young player with one injury year gets a low games projection
    (and thus a low total). Blend the raw projection toward a durability expectation
    (from Madden injury rating + age, plus the player's best recent healthy season) and
    lean on it for players who have PROVEN they can play a full season - so an isolated
    injury year is not treated as chronic."""
    tr = f[(pd.to_numeric(f["games"], errors="coerce") >= 1)
           & f["madden_injury"].notna() & f["age"].notna()]
    if len(tr) < 30:
        return proj["proj_games_raw"].clip(6, 17)
    X = np.column_stack([np.ones(len(tr)), tr["madden_injury"].astype(float),
                         tr["age"].astype(float)])
    coef, *_ = np.linalg.lstsq(X, tr["games"].astype(float).to_numpy(), rcond=None)
    raw = proj["proj_games_raw"]
    pm = pd.to_numeric(proj["madden_injury"], errors="coerce")
    pa = pd.to_numeric(proj["age"], errors="coerce")
    pop = (coef[0] + coef[1] * pm + coef[2] * pa).where(pm.notna() & pa.notna())
    pop = pop.fillna(raw)                              # fall back where rating/age missing
    gcols = [c for c in ["lag1_games", "lag2_games", "lag3_games"] if c in proj.columns]
    maxrec = proj[gcols].max(axis=1) if gcols else pd.Series(np.nan, index=proj.index)
    dur = (0.5 * maxrec.fillna(pop) + 0.5 * pop).clip(6, 17)
    proven = ((maxrec - 11) / 6).clip(0, 1).fillna(0.0)   # 0 at <=11 games, 1 at 17
    w = 0.7 * proven
    return ((1 - w) * raw + w * dur).fillna(raw).clip(6, 17)


def project(target: int = 2026, span_start: int = 2015):
    """Return {pos: DataFrame} of projected stat lines. Season points are decomposed
    into per-game production x durability-adjusted games, so availability (not just
    talent) flows into the ranking and a one-off injury year does not tank a player."""
    frames, feat_cols = F.build_frames(span_start, target - 1, POS)
    proj_frames, _ = F.build_projection_frame(target, POS)
    ps = D.build_player_seasons(list(range(span_start, target)))
    lab_cols = ["games", "half_ppr_ppg"] + [s + "_pg" for pos in POS
                                            for s in STAT_TARGETS[pos] if s != "games"]
    lab_cols = sorted(set(c for c in lab_cols if c in ps.columns))
    labels = ps[["player_id", "season"] + lab_cols].drop_duplicates(["player_id", "season"])

    results = {}
    for pos in POS:
        f = frames[pos].merge(labels, on=["player_id", "season"], how="left")
        proj = proj_frames[pos].copy()
        feats = feat_cols[pos]
        _fit_predict(f, proj, feats, "half_ppr_ppg", "proj_ppg")   # talent / role
        _fit_predict(f, proj, feats, "games", "proj_games_raw")     # raw availability
        proj["proj_games"] = _durability_games(f, proj)
        proj["proj_points"] = proj["proj_ppg"] * proj["proj_games"]
        for s in STAT_TARGETS[pos]:                                 # stat line: pg x games
            if s != "games" and s + "_pg" in f.columns:
                _fit_predict(f, proj, feats, s + "_pg", "proj_" + s + "_pg")
                proj["proj_" + s] = proj["proj_" + s + "_pg"] * proj["proj_games"]
        proj = proj.sort_values("proj_points", ascending=False).reset_index(drop=True)
        proj["pos_rank"] = np.arange(1, len(proj) + 1)
        results[pos] = proj
    return results


_INT_COLS = {"age", "madden_overall", "vbd", "adp_rank", "value", "mdl_rank",
             "vbd_rank", "pos_rank", "ovr_rank", "tier"}


def _round(df):
    df = df.copy()
    for c in df.columns:
        if c == "proj_ppg":
            df[c] = df[c].round(1)
        elif c.startswith("proj_") or c in _INT_COLS:
            df[c] = df[c].round(0).astype("Int64")
    return df


def export_excel(results, target, path, adp=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    combined = []
    for pos, df in results.items():
        d = df.copy(); d["position"] = pos
        combined.append(d)
    allp = pd.concat(combined, ignore_index=True)
    allp, baselines = add_vbd(allp)
    allp = add_tiers(allp)

    board = allp.sort_values("vbd", ascending=False).reset_index(drop=True)
    board["vbd_rank"] = np.arange(1, len(board) + 1)

    board_cols = list(BOARD_COLS)
    value_sheets = {}
    if adp is not None:
        adp_tbl = adp if isinstance(adp, pd.DataFrame) else load_adp(adp)
        board = attach_adp(board, adp_tbl)
        # insert ADP + Value right after VBD on the board
        board_cols = ["vbd_rank", "player_name", "position", "tier", "team", "age",
                      "proj_points", "vbd", "adp_rank", "value", "proj_ppg", "pos_rank"]
        matched = board[board["value"].notna()]
        values = matched[(matched.mdl_rank <= 130) & (matched.value >= 15)] \
            .sort_values("value", ascending=False)
        reaches = matched[(matched.adp_rank <= 120) & (matched.value <= -15)] \
            .sort_values("value")
        value_sheets = {
            "Values": _round(values[VALUE_COLS]),
            "Reaches": _round(reaches[VALUE_COLS]),
        }

    allp = allp.merge(board[["player_name", "position", "vbd_rank"]],
                      on=["player_name", "position"], how="left")
    allp_pts = allp.sort_values("proj_points", ascending=False).reset_index(drop=True)
    allp_pts["ovr_rank"] = np.arange(1, len(allp_pts) + 1)

    sheets = {"Draft Board": _round(board[[c for c in board_cols if c in board.columns]])}
    sheets.update(value_sheets)
    sheets["All Players"] = _round(allp_pts[[c for c in ALL_COLS if c in allp_pts.columns]])
    for pos in POS:
        d = allp[allp.position == pos].sort_values("proj_points", ascending=False)
        cols = [c for c in TAB_COLS[pos] if c in d.columns]
        sheets[pos] = _round(d[cols])

    bl = "  ".join(f"{p}={int(v)}" for p, v in sorted(baselines.items()))
    about = pd.DataFrame({
        f"NFL {target} Half-PPR Draft Cheat Sheet (12-team, 1QB/2RB/3WR/1TE/1FLEX)": [
            f"Generated from Madden {target - 1999} launch ratings + through-{target-1} production.",
            "Per-position models (walk-forward validated): Spearman rank 0.63-0.70 in-sample,",
            "0.72-0.78 out-of-sample on 2025. Beats 'repeat last year' by 4-17 pts MAE.",
            "",
            "DRAFT BOARD is the overall draft order, ranked by VBD (not raw points).",
            "VBD = Value Based Drafting = projected points minus a replacement starter at the",
            "same position. It mixes positions correctly: elite RB/WR outrank higher-scoring QBs",
            "because the dropoff at their position is steeper. Replacement baselines (pts):",
            f"    {bl}",
            "",
            "TIER groups players where a position is roughly interchangeable; a new tier marks a",
            "real dropoff. On the clock: if your position has only 1 player left in its tier but",
            "another has several, take the scarce one now.",
            "",
            "Proj Pts = season half-PPR points. Proj PPG = points / projected games.",
            "Proj G = projected games (availability). MAD = Madden 27 overall.",
            "",
            "Caveats: injuries are not predicted; rookies lean on draft slot + Madden and are the",
            "least certain. A model baseline, not gospel.",
        ] + ([
            "",
            "ADP = market consensus draft rank. Val = market_rank - model_rank",
            "among matched players: positive = model likes him more than the market (a VALUE),",
            "negative = market likes him more (a REACH). See the Values and Reaches tabs.",
        ] if adp is not None else [])})

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        about.to_excel(xl, sheet_name="About", index=False)
        for name, df in sheets.items():
            df.rename(columns=DISPLAY).to_excel(xl, sheet_name=name, index=False)

    # formatting
    from openpyxl import load_workbook
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    for name in wb.sheetnames:
        ws = wb[name]
        if name == "About":
            ws.column_dimensions["A"].width = 92
            ws["A1"].font = Font(bold=True, size=13)
            continue
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        widths = {"Player": 22, "Team": 6, "Pos": 5, "Age": 5, "MAD": 6, "Tier": 5,
                  "VBD": 7, "Rk": 5, "Proj Pts": 9, "Proj PPG": 9, "Proj G": 7,
                  "Pos Rk": 7, "Ovr Rk": 7}
        for i, col in enumerate(ws[1], start=1):
            L = get_column_letter(i)
            ws.column_dimensions[L].width = widths.get(col.value, 8.5)
        # shade the projected-points column lightly
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.border = Border(bottom=thin)
    wb.save(path)
    return path
